"""Download and prepare large, reproducible NICE/ICE QSAR training datasets.

The official ICE workbooks contain endpoint records and a separate chemical
identity table.  This script joins them by stable identifiers, applies the
same conservative endpoint rules used by RalphGuard's NICE evidence service,
removes exact-identity label conflicts, and writes the four CSV files consumed
by candidate-v2 training.

It intentionally excludes predictions, mixtures, unsupported assays,
ambiguous bounds/units, and records without a defined molecular structure.
Direct human/animal sensitization calls are accepted; in-vitro sensitization
calls require concordance across at least two biological key events and receive
lower sample weight. No label is copied from a RalphGuard or CATMoS model.

Typical use inside the reproducible scientific image::

    python scripts/prepare_ice_bulk_training.py --download --min-total 10000
    python scripts/check_training_integrity.py --require-all
    python scripts/train_candidate_v2.py --validation-profile large
"""
from __future__ import annotations

import argparse
from collections import Counter, defaultdict
import csv
from dataclasses import dataclass
from datetime import date, datetime, timezone
import hashlib
import importlib.util
import json
from pathlib import Path
import re
import shutil
import sys
from typing import Any, Iterable
from urllib.request import Request, urlopen


BASE = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE_DIR = BASE / "data" / "raw" / "ice_bulk_sources"
DEFAULT_OUTPUT_DIR = BASE / "data" / "raw"
DEFAULT_STAGING_DIR = BASE / "data" / "staging"
DEFAULT_MANIFEST = DEFAULT_OUTPUT_DIR / "dataset_manifest.json"
DEFAULT_SUMMARY = DEFAULT_STAGING_DIR / "ice_bulk_preparation_summary.json"
DEFAULT_REVIEW_QUEUE = DEFAULT_STAGING_DIR / "ice_bulk_review_queue.csv"


@dataclass(frozen=True)
class BulkDataset:
    endpoint: str
    source_filename: str
    output_filename: str
    source_url: str
    assay_description: str
    positive_definition: str
    negative_definition: str


DATASETS = {
    "skin": BulkDataset(
        endpoint="skin",
        source_filename="skin_irritation.xlsx",
        output_filename="skin_irritation.csv",
        source_url="https://ice.ntp.niehs.nih.gov/downloads/DataonICE/skin_irritation.xlsx",
        assay_description=(
            "NICE/ICE in-vivo rabbit/human evidence plus OECD TG 439 reconstructed-human-"
            "epidermis irritation calls and positive TG 431/430/435 corrosion calls"
        ),
        positive_definition="explicit irritant, corrosive, or GHS skin category 1/2 classification",
        negative_definition="explicit not-classified or non-irritant skin classification",
    ),
    "eye": BulkDataset(
        endpoint="eye",
        source_filename="eye_irritation.xlsx",
        output_filename="eye_irritation.csv",
        source_url="https://ice.ntp.niehs.nih.gov/downloads/DataonICE/eye_irritation.xlsx",
        assay_description=(
            "NICE/ICE in-vivo rabbit evidence plus OECD TG 494 Vitrigel explicit "
            "No-Category calls"
        ),
        positive_definition="explicit irritant, corrosive, or GHS eye category 1/2 classification",
        negative_definition="explicit not-classified or non-irritant eye classification",
    ),
    "sens": BulkDataset(
        endpoint="sens",
        source_filename="skin_sensitization.xlsx",
        output_filename="skin_sensitization.csv",
        source_url="https://ice.ntp.niehs.nih.gov/downloads/DataonICE/skin_sensitization.xlsx",
        assay_description="NICE/ICE human/animal direct sensitization calls plus cross-key-event in-vitro consensus",
        positive_definition="explicit sensitizer/Active classification, reported LLNA EC3 or SI >= 3, or concordant calls across >=2 in-vitro key events",
        negative_definition="explicit non-sensitizer/Inactive classification or concordant calls across >=2 in-vitro key events; a single SI < 3 is insufficient",
    ),
    "acute": BulkDataset(
        endpoint="acute",
        source_filename="acute_oral.xlsx",
        output_filename="acute_oral_toxicity.csv",
        source_url="https://ice.ntp.niehs.nih.gov/downloads/DataonICE/acute_oral.xlsx",
        assay_description="NICE/ICE in-vivo rat acute oral toxicity records",
        positive_definition="interpretable oral LD50 <= 2000 mg/kg or explicit GHS acute category 1-4",
        negative_definition="interpretable oral LD50 > 2000 mg/kg or explicit GHS acute category 5",
    ),
}


# OECD skin-sensitization defined approaches combine information from distinct
# biological key events.  A single in-vitro assay call is therefore not used as
# a training label.  The bulk-only aggregator below requires concordant calls
# from at least two different key events and assigns them a lower sample weight.
SENSITIZATION_IN_VITRO_KEY_EVENTS = {
    "DPRA": "key_event_1_protein_binding",
    "KeratinoSens": "key_event_2_keratinocyte_activation",
    "LuSens": "key_event_2_keratinocyte_activation",
    "h-CLAT": "key_event_3_dendritic_cell_activation",
    "U-SENS": "key_event_3_dendritic_cell_activation",
    "mMUSST": "key_event_3_dendritic_cell_activation",
}
SENSITIZATION_IN_VITRO_BY_NORMALIZED_NAME = {
    normalize_name.casefold(): normalize_name
    for normalize_name in SENSITIZATION_IN_VITRO_KEY_EVENTS
}

# These method names are matched exactly to the assay names in the official ICE
# workbooks.  OECD TG 439 can identify both UN-GHS Category 2 and No Category,
# so an explicit Active/Inactive ``Call`` is eligible as a lower-weight binary
# irritation label.  A negative corrosion call only means "not corrosive" and
# must not be converted to "not irritating"; TG 431/430/435 therefore supply
# positive hazard votes only in this binary endpoint.
SKIN_IN_VITRO_IRRITATION_ASSAYS = {
    "EpiDerm Irritation",
    "EpiSkin Irritation",
    "LabCyte EPI-MODEL24 Irritation",
}
SKIN_IN_VITRO_CORROSION_ASSAYS = {
    "EpiDerm Corrosion",
    "EpiSkin Corrosion",
    "SkinEthic Corrosion",
    "LabCyte Corrosion",
    "TER Corrosion",
    "Corrositex",
}
SKIN_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME = {
    name.casefold(): name
    for name in SKIN_IN_VITRO_IRRITATION_ASSAYS | SKIN_IN_VITRO_CORROSION_ASSAYS
}
EYE_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME = {"vitrigel": "Vitrigel"}


IDENTITY_FIELDS = {
    "dtxsid": ("dtxsid", "dsstox_substance_id", "dsstoxsid"),
    "inchikey": ("inchikey", "inchi_key"),
    "cid": ("cid", "pubchem_cid"),
    "casrn": ("casrn", "cas_rn", "cas_number", "cas"),
}
SMILES_FIELDS = (
    "smiles",
    "canonical_smiles",
    "qsar_ready_smiles",
    "qsar_readysmiles",
    "isomeric_smiles",
)
NAME_FIELDS = ("preferred_name", "chemical_name", "substance_name", "name")
HEADER_HINTS = {
    *SMILES_FIELDS,
    "dtxsid",
    "inchikey",
    "endpoint",
    "response",
    "assay",
    "record_id",
    "data_type",
}


def _load_nice_evidence_module():
    """Load the pure-stdlib harmonizer without importing the full backend app."""
    path = BASE / "backend" / "app" / "services" / "nice_evidence.py"
    spec = importlib.util.spec_from_file_location("ralphguard_nice_evidence", path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"cannot load NICE harmonizer from {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def _unique_headers(values: Iterable[Any]) -> list[str]:
    counts: Counter[str] = Counter()
    headers: list[str] = []
    for index, value in enumerate(values, start=1):
        base = normalize_header(value) or f"column_{index}"
        counts[base] += 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def iter_sheet_records(sheet) -> Iterable[tuple[int, dict[str, Any]]]:
    """Yield row dictionaries while tolerating title/notes before the header."""
    rows = sheet.iter_rows(values_only=True)
    headers: list[str] | None = None
    header_row = 0
    for row_number, values in enumerate(rows, start=1):
        candidate = _unique_headers(values)
        recognized = sum(header in HEADER_HINTS for header in candidate)
        nonempty = sum(value not in (None, "") for value in values)
        if recognized >= 1 and nonempty >= 3:
            headers = candidate
            header_row = row_number
            break
    if headers is None:
        return
    for row_number, values in enumerate(rows, start=header_row + 1):
        if not any(value not in (None, "") for value in values):
            continue
        yield row_number, {
            headers[index]: value
            for index, value in enumerate(values[: len(headers)])
            if value not in (None, "")
        }


def field(record: dict[str, Any], *aliases: str) -> Any:
    for alias in aliases:
        key = normalize_header(alias)
        if key in record and record[key] not in (None, ""):
            return record[key]
        # Excel sometimes contains duplicate headings such as "CAS RN...2".
        for actual_key, value in record.items():
            if actual_key.startswith(f"{key}_") and value not in (None, ""):
                return value
    return None


def compact_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def identity_aliases(record: dict[str, Any]) -> list[str]:
    aliases: list[str] = []
    for kind, names in IDENTITY_FIELDS.items():
        value = compact_text(field(record, *names))
        if value:
            aliases.append(f"{kind}:{value.casefold()}")
    return aliases


def identity_from_record(record: dict[str, Any]) -> dict[str, str] | None:
    smiles = compact_text(field(record, *SMILES_FIELDS))
    if not smiles:
        return None
    return {
        "smiles": smiles,
        "inchikey": compact_text(field(record, *IDENTITY_FIELDS["inchikey"])),
        "dtxsid": compact_text(field(record, *IDENTITY_FIELDS["dtxsid"])),
        "casrn": compact_text(field(record, *IDENTITY_FIELDS["casrn"])),
        "cid": compact_text(field(record, *IDENTITY_FIELDS["cid"])),
        "name": compact_text(field(record, *NAME_FIELDS)),
    }


def build_identity_lookup(workbook) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    candidate_sheets = [
        sheet for sheet in workbook.worksheets
        if "chemical" in normalize_header(sheet.title)
    ]
    # Some ICE releases put structures directly in a non-obviously named sheet.
    if not candidate_sheets:
        candidate_sheets = list(workbook.worksheets)
    for sheet in candidate_sheets:
        for _, record in iter_sheet_records(sheet):
            identity = identity_from_record(record)
            if identity is None:
                continue
            for alias in identity_aliases(record):
                lookup.setdefault(alias, identity)
    return lookup


def resolve_identity(
    record: dict[str, Any],
    lookup: dict[str, dict[str, str]],
) -> dict[str, str] | None:
    direct = identity_from_record(record)
    if direct is not None:
        return direct
    for alias in identity_aliases(record):
        if alias in lookup:
            return lookup[alias]
    return None


def normalized_sheet_name(title: str) -> str:
    return normalize_header(title).replace("_", " ")


def select_assay(dataset: BulkDataset, record: dict[str, Any], sheet_name: str) -> tuple[str | None, str]:
    mixture = compact_text(field(record, "mixture", "substance_type")).casefold()
    if mixture and mixture not in {"chemical", "single chemical", "single substance"}:
        return None, "mixture_or_formulation"

    assay_text = compact_text(field(record, "assay", "assay_name", "method"))
    assay_name = assay_text.casefold()
    species = compact_text(field(record, "species", "test_species")).casefold()
    data_type = compact_text(field(record, "data_type", "datatype", "study_type")).casefold()
    combined = " | ".join(
        [
            assay_text,
            data_type,
            sheet_name,
        ]
    ).casefold()

    if dataset.endpoint == "sens":
        if "human repeat insult patch" in assay_name:
            return "Human Repeat Insult Patch Test", "accepted_human_assay"
        if "human maximization" in assay_name or ("maximization" in assay_name and "human" in species):
            return "Human Maximization Test", "accepted_human_assay"
        if any(token in assay_name for token in ("llna", "local lymph node")):
            return "Murine Local Lymph Node Assay (LLNA)", "accepted_assay"
        if any(token in assay_name for token in ("guinea", "buehler")) or (
            "maximization" in assay_name and "guinea" in species
        ):
            return "Guinea Pig Maximization/Buehler", "accepted_assay"
        in_vitro_assay = SENSITIZATION_IN_VITRO_BY_NORMALIZED_NAME.get(assay_name)
        if in_vitro_assay is not None:
            return in_vitro_assay, "accepted_in_vitro_consensus_candidate"
        if data_type and "in vivo" not in data_type and "animal" not in data_type:
            return None, "unsupported_sensitization_in_vitro_assay"
        return None, "unsupported_sensitization_assay"

    if dataset.endpoint == "skin":
        in_vitro_assay = SKIN_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME.get(assay_name)
        if in_vitro_assay is not None:
            return in_vitro_assay, "accepted_oecd_in_vitro_assay"

    if dataset.endpoint == "eye":
        in_vitro_assay = EYE_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME.get(assay_name)
        if in_vitro_assay is not None:
            return in_vitro_assay, "accepted_oecd_in_vitro_assay"

    if data_type and "in vivo" not in data_type and "animal" not in data_type:
        return None, "not_in_vivo"

    if dataset.endpoint == "acute":
        if species and "rat" not in species:
            return None, "non_rat_acute_record"
        if "catmos" in combined or "prediction" in combined or "in silico" in combined:
            return None, "prediction_excluded"
        return "Rat Acute Oral Toxicity", "accepted_assay"

    if dataset.endpoint == "skin":
        if any(token in combined for token in ("rabbit", "draize", "in vivo", "invivo", "animal")):
            return "Rabbit Draize Skin Irritation/Corrosion Test", "accepted_assay"
        return None, "unsupported_skin_assay"

    if dataset.endpoint == "eye":
        if any(token in combined for token in ("rabbit", "draize", "in vivo", "invivo", "animal")):
            return "Rabbit Draize Eye Irritation/Corrosion Test", "accepted_assay"
        return None, "unsupported_eye_assay"

    return None, "unsupported_assay"


def evidence_record(
    dataset: BulkDataset,
    record: dict[str, Any],
    identity: dict[str, str],
    sheet_name: str,
    row_number: int,
    assay: str,
) -> dict[str, Any]:
    endpoint_value = field(record, "endpoint", "endpoint_name", "effect", "parameter")
    response = field(record, "response", "value", "result", "endpoint_value", "classification")
    unit = field(record, "response_unit", "unit", "units")
    record_id = compact_text(field(record, "record_id", "recordid", "study_id"))
    evidence_id = record_id or f"{dataset.source_filename}:{sheet_name}:{row_number}"
    raw_record = {key: compact_text(value) for key, value in record.items()}
    return {
        "ralphguard_endpoint": dataset.endpoint,
        "assay": assay,
        "ice_endpoint": compact_text(endpoint_value),
        "ice_value": compact_text(response),
        "ice_unit": compact_text(unit),
        "ice_dtxsid": identity.get("dtxsid", ""),
        "ice_casrn": identity.get("casrn", ""),
        "evidence_id": evidence_id,
        "source_sheet": sheet_name,
        "source_row": row_number,
        "identity": identity,
        "raw_record": raw_record,
    }


def molecular_group_key(identity: dict[str, str]) -> str:
    if identity.get("inchikey"):
        return f"inchikey:{identity['inchikey'].casefold()}"
    if identity.get("dtxsid"):
        return f"dtxsid:{identity['dtxsid'].casefold()}"
    return f"smiles:{identity['smiles']}"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _in_vitro_sensitization_vote(record: dict[str, Any]) -> dict[str, Any]:
    """Return an assay-level proxy vote only for an explicit ICE Call field."""
    endpoint_text = compact_text(record.get("ice_endpoint")).casefold()
    value_text = compact_text(record.get("ice_value")).casefold()
    assay = compact_text(record.get("assay"))
    if endpoint_text != "call" or value_text not in {"active", "inactive"}:
        return {
            **record,
            "candidate_label": None,
            "mapping_status": "proxy_not_used",
            "mapping_rule": "in_vitro_non_call_excluded",
            "mapping_reason": "only explicit Active/Inactive assay calls enter cross-key-event consensus",
        }
    label = 1 if value_text == "active" else 0
    return {
        **record,
        "candidate_label": label,
        "mapping_status": "supportive_in_vitro_proxy",
        "mapping_rule": f"in_vitro_{normalize_header(assay)}_{value_text}",
        "mapping_reason": "explicit in-vitro assay call; requires concordant evidence from another key event",
    }


def _explicit_active_inactive_call(record: dict[str, Any]) -> int | None:
    """Parse only a chemical-level ICE Call, never a numeric observation."""
    endpoint_text = compact_text(record.get("ice_endpoint")).casefold()
    value_text = compact_text(record.get("ice_value")).casefold()
    if endpoint_text != "call" or value_text not in {"active", "inactive"}:
        return None
    return 1 if value_text == "active" else 0


def _aggregate_skin_in_vitro(records: list[dict[str, Any]]) -> dict[str, Any]:
    irritation_labels: set[int] = set()
    corrosion_positive = False
    mapped_records: list[dict[str, Any]] = []
    for record in records:
        assay = compact_text(record.get("assay"))
        label = _explicit_active_inactive_call(record)
        mapped = dict(record)
        if label is None:
            mapped.update({
                "candidate_label": None,
                "mapping_status": "proxy_not_used",
                "mapping_rule": "in_vitro_non_call_excluded",
                "mapping_reason": "only explicit Active/Inactive Call rows are eligible",
            })
        elif assay in SKIN_IN_VITRO_IRRITATION_ASSAYS:
            irritation_labels.add(label)
            mapped.update({
                "candidate_label": label,
                "mapping_status": "supportive_oecd_in_vitro",
                "mapping_rule": f"tg439_{normalize_header(assay)}_{'active' if label else 'inactive'}",
                "mapping_reason": "explicit OECD TG 439 irritation classification call",
            })
        elif assay in SKIN_IN_VITRO_CORROSION_ASSAYS and label == 1:
            corrosion_positive = True
            mapped.update({
                "candidate_label": 1,
                "mapping_status": "supportive_oecd_in_vitro",
                "mapping_rule": f"skin_corrosion_{normalize_header(assay)}_active",
                "mapping_reason": "explicit positive skin-corrosion call establishes skin hazard",
            })
        else:
            mapped.update({
                "candidate_label": None,
                "mapping_status": "supportive_negative_only",
                "mapping_rule": "skin_corrosion_inactive_not_irritation_negative",
                "mapping_reason": "not-corrosive does not establish not-irritating",
            })
        mapped_records.append(mapped)

    if corrosion_positive:
        label, status, reason = 1, "candidate_requires_review", "positive OECD skin-corrosion call"
    elif irritation_labels == {1}:
        label, status, reason = 1, "candidate_requires_review", "concordant OECD TG 439 active call(s)"
    elif irritation_labels == {0}:
        label, status, reason = 0, "candidate_requires_review", "concordant OECD TG 439 inactive call(s)"
    elif irritation_labels == {0, 1}:
        label, status, reason = None, "conflict_review_required", "OECD TG 439 methods contain conflicting calls"
    else:
        label, status, reason = None, "review_required", "no eligible skin-irritation call or positive corrosion call"
    return {
        "candidate_label": label,
        "mapping_status": status,
        "mapping_reason": reason,
        "record_count": len(records),
        "mapped_record_count": sum(item.get("candidate_label") in {0, 1} for item in mapped_records),
        "records": mapped_records,
        "label_quality": "ice_curated_oecd_in_vitro_skin",
        "sample_weight": 0.7 if label in {0, 1} else 0.0,
    }


def _aggregate_eye_in_vitro(records: list[dict[str, Any]]) -> dict[str, Any]:
    """Use TG 494 only for its validated No-Category (Inactive) purpose."""
    calls: set[int] = set()
    mapped_records: list[dict[str, Any]] = []
    for record in records:
        label = _explicit_active_inactive_call(record)
        mapped = dict(record)
        if label == 0:
            calls.add(0)
            mapped.update({
                "candidate_label": 0,
                "mapping_status": "supportive_oecd_in_vitro",
                "mapping_rule": "tg494_vitrigel_inactive_no_category",
                "mapping_reason": "Vitrigel inactive call identifies UN-GHS No Category within its applicability domain",
            })
        elif label == 1:
            calls.add(1)
            mapped.update({
                "candidate_label": None,
                "mapping_status": "supportive_positive_only",
                "mapping_rule": "tg494_vitrigel_active_requires_follow_up",
                "mapping_reason": "an active Vitrigel call cannot by itself assign an eye-hazard category",
            })
        else:
            mapped.update({
                "candidate_label": None,
                "mapping_status": "proxy_not_used",
                "mapping_rule": "vitrigel_non_call_excluded",
                "mapping_reason": "only the explicit Inactive Call is eligible",
            })
        mapped_records.append(mapped)

    if calls == {0}:
        label, status, reason = 0, "candidate_requires_review", "OECD TG 494 Vitrigel No-Category call"
    elif calls == {0, 1}:
        label, status, reason = None, "conflict_review_required", "Vitrigel records contain conflicting calls"
    else:
        label, status, reason = None, "review_required", "TG 494 does not turn an Active call into a definitive positive label"
    return {
        "candidate_label": label,
        "mapping_status": status,
        "mapping_reason": reason,
        "record_count": len(records),
        "mapped_record_count": sum(item.get("candidate_label") in {0, 1} for item in mapped_records),
        "records": mapped_records,
        "label_quality": "ice_curated_oecd_tg494_no_category",
        "sample_weight": 0.7 if label == 0 else 0.0,
    }


def aggregate_training_endpoint(
    dataset: BulkDataset,
    records: list[dict[str, Any]],
    harmonizer,
) -> dict[str, Any]:
    """Aggregate higher-tier direct evidence before conservative in-vitro evidence."""
    if dataset.endpoint in {"skin", "eye"}:
        if dataset.endpoint == "skin":
            in_vitro_assays = set(SKIN_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME.values())
        else:
            in_vitro_assays = set(EYE_IN_VITRO_ASSAYS_BY_NORMALIZED_NAME.values())
        direct_records = [record for record in records if record.get("assay") not in in_vitro_assays]
        in_vitro_records = [record for record in records if record.get("assay") in in_vitro_assays]
        direct_result = harmonizer.aggregate_endpoint(direct_records)
        direct_label = direct_result.get("candidate_label")
        direct_status = compact_text(direct_result.get("mapping_status"))
        if direct_label in {0, 1}:
            return {
                **direct_result,
                "label_quality": "ice_curated_direct_human_or_animal_experimental",
                "sample_weight": 1.0,
            }
        if direct_status == "conflict_review_required":
            return {
                **direct_result,
                "label_quality": "excluded_direct_evidence_conflict",
                "sample_weight": 0.0,
            }
        if not in_vitro_records:
            return {
                **direct_result,
                "label_quality": "excluded_unmapped_direct_evidence",
                "sample_weight": 0.0,
            }
        return (
            _aggregate_skin_in_vitro(in_vitro_records)
            if dataset.endpoint == "skin"
            else _aggregate_eye_in_vitro(in_vitro_records)
        )

    if dataset.endpoint != "sens":
        result = harmonizer.aggregate_endpoint(records)
        result["label_quality"] = "ice_curated_rule_mapped_experimental"
        result["sample_weight"] = 1.0
        return result

    direct_records = [
        record
        for record in records
        if record.get("assay") not in SENSITIZATION_IN_VITRO_KEY_EVENTS
    ]
    proxy_records = [
        record
        for record in records
        if record.get("assay") in SENSITIZATION_IN_VITRO_KEY_EVENTS
    ]
    direct_result = harmonizer.aggregate_endpoint(direct_records)
    proxy_mapped = [_in_vitro_sensitization_vote(record) for record in proxy_records]
    mapped_records = [*(direct_result.get("records") or []), *proxy_mapped]

    direct_label = direct_result.get("candidate_label")
    direct_status = compact_text(direct_result.get("mapping_status"))
    if direct_label in {0, 1}:
        return {
            **direct_result,
            "records": mapped_records,
            "label_quality": "ice_curated_direct_human_or_animal_experimental",
            "sample_weight": 1.0,
        }
    if direct_status == "conflict_review_required":
        return {
            **direct_result,
            "records": mapped_records,
            "label_quality": "excluded_direct_evidence_conflict",
            "sample_weight": 0.0,
        }

    assay_votes: dict[str, int] = {}
    conflicting_assays: set[str] = set()
    by_assay: dict[str, set[int]] = defaultdict(set)
    for record in proxy_mapped:
        label = record.get("candidate_label")
        if label in {0, 1}:
            by_assay[compact_text(record.get("assay"))].add(int(label))
    for assay, labels in by_assay.items():
        if len(labels) == 1:
            assay_votes[assay] = next(iter(labels))
        elif labels:
            conflicting_assays.add(assay)

    by_key_event: dict[str, set[int]] = defaultdict(set)
    for assay, label in assay_votes.items():
        by_key_event[SENSITIZATION_IN_VITRO_KEY_EVENTS[assay]].add(label)
    key_event_votes = {
        key_event: next(iter(labels))
        for key_event, labels in by_key_event.items()
        if len(labels) == 1
    }
    labels = set(key_event_votes.values())
    if len(key_event_votes) >= 2 and len(labels) == 1 and not conflicting_assays:
        label = next(iter(labels))
        consensus_record = {
            "candidate_label": label,
            "mapping_status": "candidate_requires_review",
            "mapping_rule": "in_vitro_cross_key_event_consensus",
            "mapping_reason": "at least two distinct sensitization key events have concordant explicit assay calls",
            "assay": "+".join(sorted(key_event_votes)),
        }
        return {
            "candidate_label": label,
            "mapping_status": "candidate_requires_review",
            "mapping_reason": consensus_record["mapping_reason"],
            "record_count": len(records),
            "mapped_record_count": sum(item.get("candidate_label") in {0, 1} for item in mapped_records),
            "records": [*mapped_records, consensus_record],
            "label_quality": "ice_curated_in_vitro_cross_key_event_consensus",
            "sample_weight": 0.7,
        }

    return {
        "candidate_label": None,
        "mapping_status": "review_required",
        "mapping_reason": "no direct call and fewer than two concordant sensitization key events",
        "record_count": len(records),
        "mapped_record_count": sum(item.get("candidate_label") in {0, 1} for item in mapped_records),
        "records": mapped_records,
        "label_quality": "excluded_insufficient_in_vitro_consensus",
        "sample_weight": 0.0,
    }


def download_file(url: str, destination: Path, *, force: bool = False) -> None:
    if destination.exists() and not force:
        print(f"source exists: {destination}", flush=True)
        return
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_suffix(destination.suffix + ".part")
    request = Request(url, headers={"User-Agent": "RalphGuard-NSC2026-data-prep/1.0"})
    print(f"downloading {url}", flush=True)
    try:
        with urlopen(request, timeout=120) as response, temporary.open("wb") as handle:
            shutil.copyfileobj(response, handle, length=1024 * 1024)
        temporary.replace(destination)
    finally:
        if temporary.exists():
            temporary.unlink()


def prepare_dataset(
    dataset: BulkDataset,
    source_path: Path,
    output_path: Path,
    harmonizer,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required; install scientific/requirements.txt") from exc

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    lookup = build_identity_lookup(workbook)
    counters: Counter[str] = Counter()
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for sheet in workbook.worksheets:
        sheet_name = normalized_sheet_name(sheet.title)
        if "chemical" in sheet_name and "data" not in sheet_name:
            continue
        for row_number, record in iter_sheet_records(sheet):
            endpoint_value = field(record, "endpoint", "endpoint_name", "effect", "parameter")
            response = field(record, "response", "value", "result", "endpoint_value", "classification")
            if endpoint_value in (None, "") and response in (None, ""):
                continue
            counters["raw_evidence_rows"] += 1
            assay, assay_status = select_assay(dataset, record, sheet_name)
            if assay is None:
                counters[assay_status] += 1
                continue
            identity = resolve_identity(record, lookup)
            if identity is None:
                counters["missing_defined_structure"] += 1
                continue
            evidence = evidence_record(
                dataset,
                record,
                identity,
                sheet_name,
                row_number,
                assay,
            )
            grouped[molecular_group_key(identity)].append(evidence)
            counters["evidence_rows_joined_to_structure"] += 1

    workbook.close()
    output_rows: list[dict[str, Any]] = []
    review_rows: list[dict[str, Any]] = []
    for identity_key, records in sorted(grouped.items()):
        result = aggregate_training_endpoint(dataset, records, harmonizer)
        label = result.get("candidate_label")
        status = compact_text(result.get("mapping_status"))
        first = records[0]
        identity = first["identity"]
        mapped_records = result.get("records") or []
        evidence_ids = sorted({compact_text(record.get("evidence_id")) for record in records})
        rules = sorted(
            {
                compact_text(record.get("mapping_rule"))
                for record in mapped_records
                if compact_text(record.get("mapping_rule"))
            }
        )
        if label not in {0, 1}:
            counters[f"group_{status or 'unmapped'}"] += 1
            review_rows.append(
                {
                    "endpoint": dataset.endpoint,
                    "identity_key": identity_key,
                    "smiles": identity.get("smiles", ""),
                    "inchikey": identity.get("inchikey", ""),
                    "name": identity.get("name", ""),
                    "mapping_status": status,
                    "mapping_reason": compact_text(result.get("mapping_reason")),
                    "record_count": len(records),
                    "evidence_ids": json.dumps(evidence_ids, ensure_ascii=False),
                    "mapping_rules": json.dumps(rules, ensure_ascii=False),
                }
            )
            continue
        counters["training_identity_groups"] += 1
        counters[f"label_{int(label)}"] += 1
        label_quality = compact_text(result.get("label_quality")) or "ice_curated_rule_mapped_experimental"
        sample_weight = float(result.get("sample_weight", 1.0))
        counters[f"quality_{label_quality}"] += 1
        output_rows.append(
            {
                "smiles": identity.get("smiles", ""),
                "label": int(label),
                "name": identity.get("name", ""),
                "inchikey": identity.get("inchikey", ""),
                "dtxsid": identity.get("dtxsid", ""),
                "casrn": identity.get("casrn", ""),
                "source": "NICEATM Integrated Chemical Environment (ICE) bulk curated experimental data",
                "source_url": dataset.source_url,
                "evidence_ids": json.dumps(evidence_ids, ensure_ascii=False),
                "mapping_rules": json.dumps(rules, ensure_ascii=False),
                "label_quality": label_quality,
                "sample_weight": sample_weight,
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "smiles",
        "label",
        "name",
        "inchikey",
        "dtxsid",
        "casrn",
        "source",
        "source_url",
        "evidence_ids",
        "mapping_rules",
        "label_quality",
        "sample_weight",
    ]
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(output_rows)

    summary = {
        "endpoint": dataset.endpoint,
        "source_file": str(source_path.relative_to(BASE)),
        "output_file": str(output_path.relative_to(BASE)),
        "source_sha256": sha256_file(source_path),
        "output_sha256": sha256_file(output_path),
        "chemical_identity_aliases": len(lookup),
        **dict(sorted(counters.items())),
    }
    return summary, review_rows


def write_review_queue(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "endpoint",
        "identity_key",
        "smiles",
        "inchikey",
        "name",
        "mapping_status",
        "mapping_reason",
        "record_count",
        "evidence_ids",
        "mapping_rules",
    ]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--download", action="store_true", help="download official ICE workbooks when absent")
    parser.add_argument("--force-download", action="store_true", help="replace existing source workbooks")
    parser.add_argument("--endpoint", choices=["all", *DATASETS], default="all")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    parser.add_argument("--summary", type=Path, default=DEFAULT_SUMMARY)
    parser.add_argument("--review-queue", type=Path, default=DEFAULT_REVIEW_QUEUE)
    parser.add_argument("--min-total", type=int, default=10_000)
    parser.add_argument(
        "--allow-under-minimum",
        action="store_true",
        help="return success even when retained unique endpoint rows are below --min-total",
    )
    args = parser.parse_args()

    selected = DATASETS if args.endpoint == "all" else {args.endpoint: DATASETS[args.endpoint]}
    for dataset in selected.values():
        source_path = args.source_dir / dataset.source_filename
        if args.download or args.force_download:
            download_file(dataset.source_url, source_path, force=args.force_download)
        if not source_path.exists():
            raise FileNotFoundError(
                f"missing {source_path}; rerun with --download or place the official ICE workbook there"
            )

    harmonizer = _load_nice_evidence_module()
    generated_at = datetime.now(timezone.utc).isoformat()
    endpoint_summaries: dict[str, dict[str, Any]] = {}
    all_review_rows: list[dict[str, Any]] = []
    manifest_datasets: dict[str, dict[str, Any]] = {}

    for endpoint, dataset in selected.items():
        source_path = args.source_dir / dataset.source_filename
        output_path = args.output_dir / dataset.output_filename
        summary, review_rows = prepare_dataset(
            dataset,
            source_path,
            output_path,
            harmonizer,
        )
        endpoint_summaries[endpoint] = summary
        all_review_rows.extend(review_rows)
        manifest_datasets[endpoint] = {
            "file": dataset.output_filename,
            "source_file": str(source_path.relative_to(BASE)),
            "source_organization": "NIH/NICEATM Integrated Chemical Environment (ICE)",
            "source_url": dataset.source_url,
            "retrieved_or_verified_at": generated_at,
            "license_or_terms": "NICE states these curated public data have no restrictions on use; retain source attribution",
            "assay_or_guideline": dataset.assay_description,
            "endpoint_definition": dataset.positive_definition,
            "negative_definition": dataset.negative_definition,
            "label_mapping": "Conservative deterministic mapping in scripts/prepare_ice_bulk_training.py using backend/app/services/nice_evidence.py",
            "raw_sha256": summary["source_sha256"],
            "prepared_sha256": summary["output_sha256"],
            "prepared_unique_rows": summary.get("training_identity_groups", 0),
            "review_note": "Mixtures, predictions, ambiguous evidence, label conflicts, single-assay in-vitro sensitization calls, and missing structures were excluded",
        }
        print(json.dumps(summary, ensure_ascii=False), flush=True)

    write_review_queue(args.review_queue, all_review_rows)
    total = sum(int(item.get("training_identity_groups", 0)) for item in endpoint_summaries.values())
    complete_all = set(selected) == set(DATASETS)
    report = {
        "generated_at": generated_at,
        "source": "NICEATM Integrated Chemical Environment bulk downloads",
        "endpoint_summaries": endpoint_summaries,
        "total_unique_endpoint_training_rows": total,
        "minimum_requested": args.min_total,
        "minimum_met": total >= args.min_total,
        "review_queue": str(args.review_queue.relative_to(BASE)),
        "review_queue_rows": len(all_review_rows),
        "label_policy": "experimental/reference evidence only; no RalphGuard or CATMoS prediction labels",
    }
    args.summary.parent.mkdir(parents=True, exist_ok=True)
    args.summary.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    if complete_all:
        manifest = {
            "schema_version": 2,
            "created_at": generated_at,
            "created_by": "scripts/prepare_ice_bulk_training.py",
            "datasets": manifest_datasets,
            "aggregate": {
                "unique_endpoint_training_rows": total,
                "minimum_requested": args.min_total,
                "minimum_met": total >= args.min_total,
            },
        }
        args.manifest.parent.mkdir(parents=True, exist_ok=True)
        args.manifest.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    print(json.dumps(report, ensure_ascii=False, indent=2))
    if total < args.min_total and not args.allow_under_minimum:
        print(
            f"ERROR: retained {total:,} unique endpoint rows, below requested {args.min_total:,}",
            file=sys.stderr,
        )
        return 3
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
